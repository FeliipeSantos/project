import io
import csv
from datetime import datetime, date
from decimal import Decimal
from typing import List
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont
try:
    from weasyprint import HTML
except (ImportError, OSError):
    HTML = None
from jinja2 import Template

from app.models.models import Transaction, Account, Investment, Category
from app.schemas.report import (
    DashboardReportResponse, CategoryReportItem, 
    AccountReportItem, CashFlowReportItem, InvestmentReportItem
)

class ReportService:
    @staticmethod
    def get_dashboard_report(db: Session, user_id: str) -> DashboardReportResponse:
        accounts = db.query(Account).filter(Account.user_id == user_id).all()
        investments = db.query(Investment).filter(Investment.user_id == user_id).all()
        transactions = db.query(Transaction).filter(Transaction.user_id == user_id).all()
        categories = db.query(Category).filter(
            (Category.user_id == user_id) | (Category.user_id == None)
        ).all()

        category_map = {c.id: c for c in categories}
        account_map = {a.id: a for a in accounts}

        # 1. Balances
        total_balance = sum(a.balance_current or Decimal("0.00") for a in accounts if not getattr(a, "ignore_in_totals", False))
        total_invested = sum(i.current_value or Decimal("0.00") for i in investments)
        total_net_worth = total_balance + total_invested

        # 2. Monthly Revenues & Expenses (Current Month)
        today = date.today()
        current_month = today.month
        current_year = today.year

        current_month_txs = [
            t for t in transactions 
            if t.date and t.date.month == current_month and t.date.year == current_year
        ]

        monthly_revenues = sum(
            t.value or Decimal("0.00") for t in current_month_txs if t.type == "REVENUE"
        )
        monthly_expenses = sum(
            t.value or Decimal("0.00") for t in current_month_txs if t.type == "EXPENSE"
        )
        monthly_savings = monthly_revenues - monthly_expenses

        # 3. Expenses by Category
        expense_by_cat = {}
        for t in current_month_txs:
            if t.type == "EXPENSE" and t.category_id:
                expense_by_cat[t.category_id] = expense_by_cat.get(t.category_id, Decimal("0.00")) + (t.value or Decimal("0.00"))

        expenses_by_category = []
        for cat_id, val in expense_by_cat.items():
            cat = category_map.get(cat_id)
            cat_name = cat.name if cat else "Sem Categoria"
            cat_color = cat.color if cat and cat.color else "#CCCCCC"
            percentage = (val / monthly_expenses * Decimal("100")) if monthly_expenses > 0 else Decimal("0.00")
            
            expenses_by_category.append(
                CategoryReportItem(
                    category_name=cat_name,
                    value=val,
                    color=cat_color,
                    percentage=percentage.quantize(Decimal("0.01"))
                )
            )
        expenses_by_category.sort(key=lambda x: x.value, reverse=True)

        # 4. Account Balances
        balance_by_account = [
            AccountReportItem(
                account_name=a.name,
                balance=a.balance_current or Decimal("0.00"),
                color=a.color
            ) for a in accounts if getattr(a, "show_in_resume", True)
        ]

        # 5. Cash Flow (current year)
        tx_by_month = {}
        for t in transactions:
            if t.date and t.date.year == current_year:
                m = t.date.month
                if m not in tx_by_month:
                    tx_by_month[m] = []
                tx_by_month[m].append(t)

        cash_flow = []
        month_names = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        for m in range(1, 13):
            month_txs = tx_by_month.get(m, [])
            rev = sum(t.value or Decimal("0.00") for t in month_txs if t.type == "REVENUE")
            exp = sum(t.value or Decimal("0.00") for t in month_txs if t.type == "EXPENSE")
            cash_flow.append(
                CashFlowReportItem(
                    month=month_names[m - 1],
                    revenues=rev,
                    expenses=exp,
                    balance=rev - exp
                )
            )

        # 6. Investments
        investments_report = []
        for i in investments:
            qty = i.quantity or Decimal("0.000000")
            avg_price = i.average_price or Decimal("0.0000")
            curr_val = i.current_value or Decimal("0.00")
            cost = qty * avg_price
            prof = curr_val - cost
            prof_percent = (prof / cost * Decimal("100")) if cost > 0 else Decimal("0.00")
            
            investments_report.append(
                InvestmentReportItem(
                    investment_name=i.name,
                    type=i.type,
                    quantity=qty,
                    average_price=avg_price,
                    cost_value=cost.quantize(Decimal("0.01")),
                    current_value=curr_val,
                    profitability=prof.quantize(Decimal("0.01")),
                    profitability_percentage=prof_percent.quantize(Decimal("0.01"))
                )
            )

        return DashboardReportResponse(
            total_balance=total_balance,
            total_invested=total_invested,
            total_net_worth=total_net_worth,
            monthly_revenues=monthly_revenues,
            monthly_expenses=monthly_expenses,
            monthly_savings=monthly_savings,
            expenses_by_category=expenses_by_category,
            balance_by_account=balance_by_account,
            cash_flow=cash_flow,
            investments=investments_report
        )

    @staticmethod
    def export_transactions_csv(db: Session, user_id: str) -> str:
        transactions = db.query(Transaction).filter(Transaction.user_id == user_id).all()
        categories = db.query(Category).filter(
            (Category.user_id == user_id) | (Category.user_id == None)
        ).all()
        accounts = db.query(Account).filter(Account.user_id == user_id).all()

        category_map = {c.id: c for c in categories}
        account_map = {a.id: a for a in accounts}

        output = io.StringIO()
        writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)
        
        # Header matches Java exactly
        writer.writerow(["ID", "Data", "Descricao", "Tipo", "Valor", "Categoria", "Conta", "Notas"])

        for t in transactions:
            cat = category_map.get(t.category_id) if t.category_id else None
            acc = account_map.get(t.account_id) if t.account_id else None
            
            writer.writerow([
                str(t.id),
                t.date.strftime("%Y-%m-%d") if t.date else "",
                t.description or "",
                t.type or "",
                str(t.value),
                cat.name if cat else "",
                acc.name if acc else "",
                t.notes or ""
            ])
            
        return output.getvalue()

    @staticmethod
    def export_transactions_excel(db: Session, user_id: str) -> bytes:
        transactions = db.query(Transaction).filter(Transaction.user_id == user_id).all()
        accounts = db.query(Account).filter(Account.user_id == user_id).all()
        investments = db.query(Investment).filter(Investment.user_id == user_id).all()
        categories = db.query(Category).filter(
            (Category.user_id == user_id) | (Category.user_id == None)
        ).all()

        category_map = {c.id: c for c in categories}
        account_map = {a.id: a for a in accounts}

        wb = Workbook()
        
        # Sheet 1: Transactions
        ws_tx = wb.active
        ws_tx.title = "Transações"
        
        tx_cols = ["Data", "Descrição", "Tipo", "Valor", "Categoria", "Conta", "Notas"]
        ws_tx.append(tx_cols)
        
        # Apply Bold Header style
        for col_num in range(1, len(tx_cols) + 1):
            cell = ws_tx.cell(row=1, column=col_num)
            cell.font = XLFont(bold=True)

        for t in transactions:
            cat = category_map.get(t.category_id) if t.category_id else None
            acc = account_map.get(t.account_id) if t.account_id else None
            ws_tx.append([
                t.date.strftime("%Y-%m-%d") if t.date else "",
                t.description or "",
                t.type or "",
                float(t.value) if t.value is not None else 0.0,
                cat.name if cat else "",
                acc.name if acc else "",
                t.notes or ""
            ])

        # Sheet 2: Accounts
        ws_acc = wb.create_sheet(title="Contas")
        acc_cols = ["Nome", "Instituição", "Tipo", "Saldo Atual"]
        ws_acc.append(acc_cols)
        
        for col_num in range(1, len(acc_cols) + 1):
            cell = ws_acc.cell(row=1, column=col_num)
            cell.font = XLFont(bold=True)

        for a in accounts:
            ws_acc.append([
                a.name or "",
                a.institution or "",
                a.type or "",
                float(a.balance_current) if a.balance_current is not None else 0.0
            ])

        # Sheet 3: Investments
        ws_inv = wb.create_sheet(title="Investimentos")
        inv_cols = ["Nome", "Tipo", "Preço Médio", "Quantidade", "Valor de Custo", "Valor Atual", "Rentabilidade", "Rentabilidade %"]
        ws_inv.append(inv_cols)
        
        for col_num in range(1, len(inv_cols) + 1):
            cell = ws_inv.cell(row=1, column=col_num)
            cell.font = XLFont(bold=True)

        for inv in investments:
            qty = inv.quantity or Decimal("0.000000")
            avg_price = inv.average_price or Decimal("0.0000")
            curr_val = inv.current_value or Decimal("0.00")
            cost = qty * avg_price
            prof = curr_val - cost
            prof_percent = (prof / cost) if cost > 0 else Decimal("0.00")

            ws_inv.append([
                inv.name or "",
                inv.type or "",
                float(avg_price),
                float(qty),
                float(cost),
                float(curr_val),
                float(prof),
                float(prof_percent)
            ])

        # Write workbook to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    @staticmethod
    def export_transactions_pdf(db: Session, user_id: str) -> bytes:
        transactions = db.query(Transaction).filter(Transaction.user_id == user_id).all()
        accounts = db.query(Account).filter(Account.user_id == user_id).all()
        categories = db.query(Category).filter(
            (Category.user_id == user_id) | (Category.user_id == None)
        ).all()

        category_map = {c.id: c for c in categories}
        account_map = {a.id: a for a in accounts}

        total_balance = sum(a.balance_current or Decimal("0.00") for a in accounts if not getattr(a, "ignore_in_totals", False))

        # Sort transactions by date descending
        transactions.sort(key=lambda x: x.date or date.min, reverse=True)

        # Prepare transaction payload matching HTML/OpenPDF columns
        tx_data = []
        for t in transactions:
            cat = category_map.get(t.category_id) if t.category_id else None
            acc = account_map.get(t.account_id) if t.account_id else None
            tx_data.append({
                "date": t.date.strftime("%Y-%m-%d") if t.date else "",
                "description": t.description or "",
                "type": t.type or "",
                "value": f"{t.value:.2f}",
                "category_name": cat.name if cat else "",
                "account_name": acc.name if acc else ""
            })

        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                @page { margin: 1.5cm; }
                body { font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; color: #333333; line-height: 1.4; }
                h1 { font-size: 18pt; font-weight: bold; margin-bottom: 4px; color: #111111; }
                .subtitle { font-size: 11pt; color: #666666; margin-bottom: 6px; }
                table { width: 100%; border-collapse: collapse; margin-top: 15px; }
                th { background-color: #e0e0e0; font-weight: bold; font-size: 9pt; text-align: left; padding: 6px 8px; border: 1px solid #cccccc; }
                td { font-size: 9pt; padding: 6px 8px; border: 1px solid #e0e0e0; }
                tr:nth-child(even) { background-color: #fafafa; }
            </style>
        </head>
        <body>
            <h1>FinControl - Relatório Financeiro</h1>
            <div class="subtitle">Gerado em: {{ today }}</div>
            <div class="subtitle">Saldo Total em Contas: R$ {{ total_balance }}</div>
            
            <table>
                <thead>
                    <tr>
                        <th>Data</th>
                        <th>Descrição</th>
                        <th>Tipo</th>
                        <th>Valor</th>
                        <th>Categoria</th>
                        <th>Conta</th>
                    </tr>
                </thead>
                <tbody>
                    {% for t in transactions %}
                    <tr>
                        <td>{{ t.date }}</td>
                        <td>{{ t.description }}</td>
                        <td>{{ t.type }}</td>
                        <td>R$ {{ t.value }}</td>
                        <td>{{ t.category_name }}</td>
                        <td>{{ t.account_name }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </body>
        </html>
        """

        today_str = date.today().strftime("%Y-%m-%d")
        formatted_balance = f"{total_balance:.2f}".replace(".", ",")
        
        template = Template(html_template)
        rendered_html = template.render(
            today=today_str,
            total_balance=formatted_balance,
            transactions=tx_data
        )

        if HTML is None:
            raise RuntimeError(
                "PDF export requires GObject dependencies which are missing on this host. "
                "Please run within Docker to use this feature."
            )
        # WeasyPrint renders standard HTML straight into a PDF byte stream
        pdf_bytes = HTML(string=rendered_html).write_pdf()
        return pdf_bytes
